from typing import List, Dict, Any, Optional, Tuple, Type, Literal
from pydantic import BaseModel, Field, create_model, ValidationError
import streamlit as st
from mistral_client import call_mistral
from pdf_extract import extract_text_from_pdf
import json, re, zipfile, io, datetime, base64
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from cover_letter_analyzer import (
    CoverLetterAnalysis, 
    analyze_cover_letter_with_ai, 
    process_cover_letter_files,
    create_cover_letter_excel_report,
    clean_json_output_cover_letter,
    get_download_link_cover_letter
)

# ---------- Base fields ----------
class Consideration(BaseModel):
    field: str
    instruction: str
    applied: bool
    impact: str

# Core sections - REMOVED experience_relevance_score (duplicate)
BASE_FIELDS: Dict[str, Tuple[Type[Any], Any]] = {
    'key_strengths': (List[str], ...),
    'key_strengths_score': (float, Field(ge=1, le=5)),
    'key_strengths_explanation': (str, ...),

    'experience_score': (float, Field(ge=1, le=5)),
    'experience_explanation': (str, ...),

    'skills_match_score': (float, Field(ge=1, le=5)),
    'skills_match_explanation': (str, ...),

    'potential_concerns': (List[str], ...),
    'recommendation': (Literal["Recommended", "Consider", "Pass"], ...),

    'candidate_name': (str, ...),
    'job_title': (str, ...),
    'department': (str, ...),

    # Use model's overall score (no recomputing)
    'overall_score': (float, Field(ge=1, le=5)),
    'overall_explanation': (str, ...),

    'custom_considerations': (List[Consideration], ...),
}

# ---------- Dynamic model helpers ----------
def take_dynamic_input(t: str, enum_vals: Optional[list] = None) -> Tuple[Type[Any], Any]:
    if t == "enum":
        if enum_vals:
            return (Literal[tuple(enum_vals)], ...)  # type: ignore
        return (str, ...)
    mapping = {"string": str, "integer": int, "float": float, "boolean": bool}
    return (mapping.get(t, str), ...)

# For each custom field X, add:
#   X (typed value), X_score: 1–5, X_explanation: str
def build_dynamic_model(custom_fields: list) -> Type[BaseModel]:
    fields = dict(BASE_FIELDS)
    for f in custom_fields:
        enum_vals = f.get('enum_vals') if f['type'] == 'enum' else None
        fields[f['name']] = take_dynamic_input(f['type'], enum_vals)
        fields[f"{f['name']}_score"] = (Optional[float], Field(default=None, ge=1, le=5))
        fields[f"{f['name']}_explanation"] = (Optional[str], Field(default=None))
    Model = create_model('EvaluationModel', **fields)
    Model.model_config = {"extra": "forbid"}
    return Model

# ---------- ZIP file processing ----------
def extract_pdfs_from_zip(zip_file) -> List[Tuple[str, bytes]]:
    """
    Extract PDF files from a ZIP archive.
    Returns a list of tuples (filename, pdf_content)
    """
    pdf_files = []
    
    try:
        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
            for file_info in zip_ref.infolist():
                # Skip directories and non-PDF files
                if file_info.is_dir() or not file_info.filename.lower().endswith('.pdf'):
                    continue
                
                # Extract the PDF content
                pdf_content = zip_ref.read(file_info.filename)
                
                # Get just the filename without path
                filename = Path(file_info.filename).name
                
                pdf_files.append((filename, pdf_content))
                
    except zipfile.BadZipFile:
        st.error("❌ Invalid ZIP file. Please upload a valid ZIP archive.")
        return []
    except Exception as e:
        st.error(f"❌ Error reading ZIP file: {str(e)}")
        return []
    
    return pdf_files

def process_uploaded_files(uploaded_files, uploaded_zip):
    """
    Process both individual PDF files and ZIP files containing PDFs.
    Returns a list of tuples (filename, file_object_or_content)
    """
    all_files = []
    
    # Process individual PDF files
    if uploaded_files:
        for file in uploaded_files:
            all_files.append((file.name, file))
    
    # Process ZIP file
    if uploaded_zip:
        st.info(f"📁 Processing ZIP file: {uploaded_zip.name}")
        pdf_files = extract_pdfs_from_zip(uploaded_zip)
        
        if pdf_files:
            st.success(f"✅ Found {len(pdf_files)} PDF files in ZIP archive")
            for filename, pdf_content in pdf_files:
                # Create a file-like object from the PDF content
                file_obj = io.BytesIO(pdf_content)
                file_obj.name = filename  # Add name attribute for compatibility
                all_files.append((filename, file_obj))
        else:
            st.warning("⚠️ No PDF files found in the ZIP archive")
    
    return all_files

# ---------- UI ----------
st.set_page_config(page_title="Resume & Cover Letter Analyzer", layout="wide")
st.title("📄 Resume & Cover Letter Analyzer with AI")

# Create tabs for different functionalities
tab1, tab2 = st.tabs(["📄 Resume Analysis", "✉️ Cover Letter Analysis"])

with tab1:
    st.header("Resume Recommender with Mistral AI")
    
    # ---------- JD inputs ----------
    job_title = st.text_input("Job Title")
    department = st.selectbox("Department", ["Engineering", "Marketing", "Design", "Data", "Other"])
    job_description = st.text_area("Job Description", height=200)

    # ---------- Core Criteria Definitions ----------
    st.subheader("Core Evaluation Criteria")

    # Initialize session state for core criteria definitions
    if 'core_criteria_defs' not in st.session_state:
        st.session_state.core_criteria_defs = {
            'key_strengths': {
                'custom': False,
                'definition': "Score key_strengths (1–5) based on job requirements"
            },
            'experience': {
                'custom': False,
                'definition': "Score experience (1–5) covering both years of experience AND relevance to this specific role"
            },
            'skills_match': {
                'custom': False,
                'definition': "Score skills_match (1–5) for technical/functional skill alignment"
            }
        }

    with st.expander("Define Core Criteria"):
        st.info("You can customize what Key Strengths, Experience, and Skills Match mean for your evaluation, or use the default definitions.")
        
        # Key Strengths definition
        st.write("**Key Strengths Definition**")
        key_strengths_custom = st.checkbox(
            "Customize Key Strengths definition", 
            value=st.session_state.core_criteria_defs['key_strengths']['custom'],
            key="key_strengths_custom"
        )
        
        if key_strengths_custom:
            key_strengths_def = st.text_area(
                "Define what Key Strengths means for this evaluation:",
                value=st.session_state.core_criteria_defs['key_strengths']['definition'],
                placeholder="Example: Key strengths should prioritize leadership abilities, technical expertise, and communication skills",
                height=80,
                key="key_strengths_def"
            )
        else:
            key_strengths_def = "Score key_strengths (1–5) based on job requirements"
        
        # Experience definition
        st.write("**Experience Definition**")
        experience_custom = st.checkbox(
            "Customize Experience definition", 
            value=st.session_state.core_criteria_defs['experience']['custom'],
            key="experience_custom"
        )
        
        if experience_custom:
            experience_def = st.text_area(
                "Define what Experience means for this evaluation:",
                value=st.session_state.core_criteria_defs['experience']['definition'],
                placeholder="Example: Experience should emphasize industry-specific background and relevant project work",
                height=80,
                key="experience_def"
            )
        else:
            experience_def = "Score experience (1–5) covering both years of experience AND relevance to this specific role"
        
        # Skills Match definition
        st.write("**Skills Match Definition**")
        skills_match_custom = st.checkbox(
            "Customize Skills Match definition", 
            value=st.session_state.core_criteria_defs['skills_match']['custom'],
            key="skills_match_custom"
        )
        
        if skills_match_custom:
            skills_match_def = st.text_area(
                "Define what Skills Match means for this evaluation:",
                value=st.session_state.core_criteria_defs['skills_match']['definition'],
                placeholder="Example: Skills match should focus on technical proficiencies listed in job description",
                height=80,
                key="skills_match_def"
            )
        else:
            skills_match_def = "Score skills_match (1–5) for technical/functional skill alignment"
        
        # Save button
        if st.button("Apply Core Criteria Definitions"):
            st.session_state.core_criteria_defs['key_strengths']['custom'] = key_strengths_custom
            st.session_state.core_criteria_defs['key_strengths']['definition'] = key_strengths_def
            
            st.session_state.core_criteria_defs['experience']['custom'] = experience_custom
            st.session_state.core_criteria_defs['experience']['definition'] = experience_def
            
            st.session_state.core_criteria_defs['skills_match']['custom'] = skills_match_custom
            st.session_state.core_criteria_defs['skills_match']['definition'] = skills_match_def
            
            st.success("✅ Core criteria definitions updated!")
    
    # ---------- Custom fields ----------
    st.subheader("Custom Evaluation Fields")
    if 'custom_fields' not in st.session_state:
        st.session_state.custom_fields = []
    
    with st.expander("Add Custom Field"):
        field_name = st.text_input("Field Name")
        field_type = st.selectbox("Field Type", ["string", "integer", "float", "boolean", "enum"])
        enum_values = []
        if field_type == "enum":
            enum_input = st.text_area("Enum Values (one per line)")
            if enum_input:
                enum_values = [v.strip() for v in enum_input.split('\n') if v.strip()]
    
        instruction = st.text_area(
            "Instruction for how to use this category in evaluation",
            placeholder=(
                "Examples:\n"
                "- If University is outside Texas, set <field>_score < 2 and explain why.\n"
                "- If publications ≥ 2, set <field>_score ≥ 4 with brief justification.\n"
            ),
            height=120
        )
    
        if st.button("Add Field"):
            if field_name:
                new_field = {
                    'name': field_name,
                    'type': field_type,
                    'enum_vals': enum_values if field_type == 'enum' else None,
                    'instruction': (instruction or "").strip() or
                                   "If relevant, set <field>_score (1–5) with one-sentence explanation referencing resume evidence."
                }
                st.session_state.custom_fields.append(new_field)
                st.success(f"Added field: {field_name}")
                st.rerun()
    
    # Display + remove
    if st.session_state.custom_fields:
        st.write("**Current Custom Fields:**")
        for i, field in enumerate(st.session_state.custom_fields):
            col1, col2, col3 = st.columns([3, 2, 1])
            with col1:
                st.write(f"• {field['name']} ({field['type']})")
                st.caption(f"Use: {field.get('instruction','')}")
            with col2:
                if field['type'] == 'enum' and field['enum_vals']:
                    st.write(f"Options: {', '.join(field['enum_vals'])}")
            with col3:
                if st.button("Remove", key=f"remove_{i}"):
                    st.session_state.custom_fields.pop(i)
                    st.rerun()
    
    # File upload section
    st.markdown("### 📁 Upload Resumes")
    
    upload_method = st.radio(
        "Choose upload method:",
        ["Upload Individual Resumes", "Upload ZIP File"],
        index=0,
        help="Individual Resumes: Select multiple PDF files. ZIP File: Upload one ZIP containing multiple PDF resumes."
    )
    
    uploaded_files = None
    uploaded_zip = None
    
    if upload_method == "Upload Individual Resumes":
        uploaded_files = st.file_uploader(
            "Select multiple PDF resume files", 
            type="pdf", 
            accept_multiple_files=True,
            key="pdf_uploader"
        )
    
    if upload_method == "Upload ZIP File":
        uploaded_zip = st.file_uploader(
            "Upload ZIP file containing PDF resumes", 
            type="zip",
            key="zip_uploader"
        )
    
    # Process all uploaded files
    all_resume_files = process_uploaded_files(uploaded_files, uploaded_zip)
    
    # Display file count summary
    if all_resume_files:
        st.info(f"📊 Ready to process {len(all_resume_files)} resume(s)")
        with st.expander("View file list"):
            for i, (filename, _) in enumerate(all_resume_files, 1):
                st.write(f"{i}. {filename}")
    
    # ---------- Prompt/schema ----------
    def schema_text(job_title: str, department: str, job_description: str, custom_fields: list) -> str:
        lines = [
            # Core - REMOVED experience_relevance (duplicate)
            '"key_strengths": ["strength1", "strength2", "strength3"],',
            '"key_strengths_score": <number 1-5>,',
            '"key_strengths_explanation": "<why this score was given for key strengths>",',
            '"experience_score": <number 1-5>,',
            '"experience_explanation": "<why this score was given for experience and relevance to role>",',
            '"skills_match_score": <number 1-5>,',
            '"skills_match_explanation": "<short, concrete rationale>",',
            '"potential_concerns": ["concern1", "concern2"],',
            '"recommendation": "<exactly one of: Recommended, Consider, Pass>",',
            '"candidate_name": "<extract from resume or use \\"Candidate\\">",',
            f'"job_title": "{job_title}",',
            f'"department": "{department}",'
        ]
    
        for f in custom_fields:
            # value
            if f["type"] == "string":
                lines.append(f'"{f["name"]}": "<string>",')
            elif f["type"] == "integer":
                lines.append(f'"{f["name"]}": <integer>,')
            elif f["type"] == "float":
                lines.append(f'"{f["name"]}": <number>,')
            elif f["type"] == "boolean":
                lines.append(f'"{f["name"]}": <true|false>,')
            elif f["type"] == "enum":
                opts = ", ".join([f'"{v}"' for v in (f.get("enum_vals") or [])]) or '"<string>"'
                lines.append(f'"{f["name"]}": <one of: {opts}>,')
            else:
                lines.append(f'"{f["name"]}": "<string>",')
            # score + explanation
            lines.append(f'"{f["name"]}_score": <number 1-5>,')
            lines.append(f'"{f["name"]}_explanation": "<short rationale tied to resume evidence>",')
    
        # Model provides overall score - no recomputing
        lines.append('"overall_score": <number 1-5>,')
        lines.append('"overall_explanation": "<1–2 sentences summarizing the key drivers from the subscores>",')
    
        lines.append('"custom_considerations": [')
        lines.append('  { "field": "<field name>", "instruction": "<the HR rule text>", "applied": <true|false>, "impact": "<what changed (e.g., university_score→1) and effect on overall>" }')
        lines.append(']')
    
        return "{\n" + "\n".join(lines) + "\n}"
    
    def build_eval_prompt(
        job_title: str,
        department: str,
        job_description: str,
        custom_fields: list,
        resume_text: str
    ) -> str:
        schema = schema_text(job_title, department, job_description, custom_fields)
        rules_payload = json.dumps(
            [{"field": f["name"], "instruction": f.get("instruction", "")} for f in custom_fields],
            ensure_ascii=False
        )
        
        # Get custom core criteria definitions if they exist, otherwise use defaults
        if 'core_criteria_defs' in st.session_state:
            key_strengths_def = st.session_state.core_criteria_defs['key_strengths']['definition']
            experience_def = st.session_state.core_criteria_defs['experience']['definition']
            skills_match_def = st.session_state.core_criteria_defs['skills_match']['definition']
        else:
            key_strengths_def = "Score key_strengths (1–5) based on job requirements"
            experience_def = "Score experience (1–5) covering both years of experience AND relevance to this specific role"
            skills_match_def = "Score skills_match (1–5) for technical/functional skill alignment"
    
        return f"""You are an expert hiring manager. Return STRICT JSON only—no prose/markdown/fences.
    
    SCORING SCALE (1-5): 5 Exceptional · 4 Strong · 3 Good · 2 Fair · 1 Poor
    
    REQUIRED JSON (exact keys/types):
    {schema}
    
    JOB:
    Title: {job_title}
    Department: {department}
    Description: {job_description}
    
    RESUME (verbatim evidence source):
    {resume_text}
    
    CATEGORY INSTRUCTIONS (authoritative; reflect ALL in custom_considerations):
    {rules_payload}
    
    EVALUATION RULES (follow ALL):
    1) {key_strengths_def}
    2) {experience_def}
    3) {skills_match_def}
    4) For EACH custom field, extract value AND provide score (1–5) AND explanation
    5) If instruction sets threshold/condition, set that field's score accordingly and note impact
    6) Calculate overall_score considering ALL individual scores (core + custom) and their relative importance
    7) If custom field has low score due to instruction, let it significantly impact overall_score
    8) overall_explanation should summarize key drivers from subscores
    9) Keep all text values concise and avoid special characters, newlines, or control characters
    10) Return ONLY the JSON object"""
    
    # ---------- Pre-Evaluation Check Functions ----------
    def validate_job_details(job_title, department, job_description):
        prompt = (
            f"Given the job title '{job_title}', department '{department}', and job description '{job_description}', "
            f"summarize the key requirements in 3-5 bullets."
        )
        return call_mistral(prompt)
    
    def validate_custom_fields(custom_fields):
        out = []
        for field in custom_fields:
            prompt = (
                f"Custom field '{field['name']}' with instruction '{field['instruction']}'. "
                f"Explain briefly how to compute a 1–5 score and give one example using resume evidence."
            )
            out.append(call_mistral(prompt))
        return out
    
    def run_pre_evaluation_checks(job_title, department, job_description, custom_fields):
        job_validation = validate_job_details(job_title, department, job_description)
        custom_field_validations = validate_custom_fields(custom_fields)
        return job_validation, custom_field_validations
    
    # ---------- Enhanced JSON sanitizer ----------
    def create_excel_report(evaluations_with_metadata):
        """
        Create a comprehensive Excel report with all evaluation results including detailed explanations
        
        Args:
            evaluations_with_metadata: List of dictionaries containing evaluation objects and metadata
            
        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Resume Evaluations"
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(color="FFFFFF", bold=True, size=10)
        score_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")  # Light green for scores
        explanation_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")  # Light orange for explanations
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create comprehensive headers
        score_headers = [
            "Candidate Name", "Job Title", "Department", "Overall Score", "Recommendation",
            "Key Strengths Score", "Experience Score", "Skills Match Score"
        ]
        
        explanation_headers = [
            "Key Strengths List", "Key Strengths Explanation", "Experience Explanation", 
            "Skills Match Explanation", "Potential Concerns", "Overall Explanation"
        ]
        
        # Add custom field headers if any
        custom_field_names = []
        custom_score_headers = []
        custom_explanation_headers = []
        
        if evaluations_with_metadata and "custom_fields" in evaluations_with_metadata[0]:
            for field in evaluations_with_metadata[0]["custom_fields"]:
                field_name = field['name']
                field_display = field_name.replace('_', ' ').title()
                custom_field_names.append(field_name)
                custom_score_headers.append(f"{field_display} Score")
                custom_explanation_headers.append(f"{field_display} Explanation")
        
        # Combine all headers
        all_score_headers = score_headers + custom_score_headers
        all_explanation_headers = explanation_headers + custom_explanation_headers
        all_headers = all_score_headers + all_explanation_headers + ["Custom Considerations", "Notes"]
        
        # Create main header row
        ws.merge_cells('A1:' + chr(ord('A') + len(all_score_headers) - 1) + '1')
        ws.cell(row=1, column=1, value="SCORES & BASIC INFO").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=1).border = border
        
        start_col = len(all_score_headers) + 1
        end_col = len(all_score_headers) + len(all_explanation_headers)
        ws.merge_cells(f'{chr(ord("A") + start_col - 1)}1:{chr(ord("A") + end_col - 1)}1')
        ws.cell(row=1, column=start_col, value="EXPLANATIONS & REASONING").font = header_font
        ws.cell(row=1, column=start_col).fill = header_fill
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=start_col).border = border
        
        # Add remaining headers for other columns
        remaining_start = len(all_score_headers) + len(all_explanation_headers) + 1
        for i, header in enumerate(["Custom Considerations", "Notes"]):
            col_num = remaining_start + i
            ws.cell(row=1, column=col_num, value=header).font = header_font
            ws.cell(row=1, column=col_num).fill = header_fill
            ws.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=col_num).border = border
        
        # Create sub-header row
        for col_num, header in enumerate(all_score_headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = subheader_font
            cell.fill = score_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        for col_num, header in enumerate(all_explanation_headers, len(all_score_headers) + 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = subheader_font
            cell.fill = explanation_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Add remaining sub-headers
        remaining_headers = ["Custom Considerations", "Notes"]
        for i, header in enumerate(remaining_headers):
            col_num = len(all_score_headers) + len(all_explanation_headers) + 1 + i
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Add evaluation data
        for row_num, eval_item in enumerate(evaluations_with_metadata, 3):
            eval_data = eval_item["evaluation"]
            
            # SCORES & BASIC INFO SECTION
            col = 1
            ws.cell(row=row_num, column=col, value=eval_data.candidate_name).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.job_title).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.department).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.overall_score).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.recommendation).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.key_strengths_score).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.experience_score).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.skills_match_score).border = border
            col += 1
            
            # Custom field scores
            for field_name in custom_field_names:
                score_attr = f"{field_name}_score"
                score_value = getattr(eval_data, score_attr, None) if hasattr(eval_data, score_attr) else None
                ws.cell(row=row_num, column=col, value=score_value).border = border
                col += 1
            
            # EXPLANATIONS & REASONING SECTION
            ws.cell(row=row_num, column=col, value=", ".join(eval_data.key_strengths)).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.key_strengths_explanation).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.experience_explanation).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.skills_match_explanation).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=", ".join(eval_data.potential_concerns)).border = border
            col += 1
            ws.cell(row=row_num, column=col, value=eval_data.overall_explanation).border = border
            col += 1
            
            # Custom field explanations
            for field_name in custom_field_names:
                explanation_attr = f"{field_name}_explanation"
                explanation_value = getattr(eval_data, explanation_attr, None) if hasattr(eval_data, explanation_attr) else None
                ws.cell(row=row_num, column=col, value=explanation_value).border = border
                col += 1
            
            # Custom considerations
            considerations_text = ""
            if hasattr(eval_data, 'custom_considerations') and eval_data.custom_considerations:
                considerations_list = []
                for item in eval_data.custom_considerations:
                    status = "APPLIED" if item.applied else "NOT APPLIED"
                    considerations_list.append(f"{item.field} → {status} | Instruction: {item.instruction} | Impact: {item.impact}")
                considerations_text = "\n".join(considerations_list)
            
            ws.cell(row=row_num, column=col, value=considerations_text).border = border
            col += 1
            
            # Notes (empty column for manual notes)
            ws.cell(row=row_num, column=col, value="").border = border
        
        # Auto-adjust column widths with better sizing (handle merged cells properly)
        def get_column_letter_simple(col_idx):
            """Simple function to convert column index to letter"""
            result = ""
            while col_idx > 0:
                col_idx -= 1
                result = chr(col_idx % 26 + ord('A')) + result
                col_idx //= 26
            return result
        
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter_simple(col_idx)
            
            # Check all cells in this column, but skip rows with merged cells in headers
            for row_idx in range(3, ws.max_row + 1):  # Start from row 3 to skip header rows
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        # Handle multi-line content
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            if len(line) > max_length:
                                max_length = len(line)
                except:
                    continue
            
            # Also check the sub-header row (row 2) for column sizing
            try:
                header_cell = ws.cell(row=2, column=col_idx)
                if header_cell.value:
                    header_length = len(str(header_cell.value))
                    if header_length > max_length:
                        max_length = header_length
            except:
                pass
            
            # Set appropriate width based on content
            if max_length < 15:
                adjusted_width = 15
            elif max_length > 60:
                adjusted_width = 60
            else:
                adjusted_width = max_length + 5
            
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights for better readability
        ws.row_dimensions[1].height = 25  # Main header
        ws.row_dimensions[2].height = 40  # Sub headers
        
        # Set text wrapping for explanation columns
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file
    
    def get_download_link(excel_file, filename):
        """
        Generate a download link for an Excel file
        
        Args:
            excel_file: BytesIO object with Excel data
            filename: Name to use for the download
            
        Returns:
            HTML string with download link
        """
        b64 = base64.b64encode(excel_file.getvalue()).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">📥 Download Excel Report</a>'
        return href
    
    def clean_json_output(raw_text: str) -> Optional[str]:
        """
        Tries to extract and sanitize a JSON object from LLM output.
        Handles control characters and malformed JSON, especially complex nested structures.
        """
        # First, try to find the JSON object with a more flexible regex
        json_match = re.search(r'\{.*\}', raw_text, re.DOTALL)
        if not json_match:
            return None
        
        s = json_match.group(0).strip()
        
        # If the raw JSON looks good already, try parsing it first
        try:
            json.loads(s)
            return s  # If it parses successfully, return as-is
        except json.JSONDecodeError:
            pass  # Continue with cleaning
        
        # Step 1: Replace smart quotes and problematic characters first
        s = s.replace('"', '"').replace('"', '"')  # Smart quotes to regular quotes
        s = s.replace(''', "'").replace(''', "'")  # Smart apostrophes
        
        # Step 2: Remove or replace control characters (but preserve structure)
        s = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', s)  # Remove control chars but keep \n, \r, \t
        
        # Step 3: Normalize whitespace without breaking structure
        s = re.sub(r'\r\n', '\n', s)  # Normalize line endings
        s = re.sub(r'\r', '\n', s)    # Convert remaining CR to LF
        
        # Step 4: Fix common JSON structural issues
        # Remove trailing commas before } or ]
        s = re.sub(r',(\s*[}\]])', r'\1', s)
        
        # Remove leading commas after { or [
        s = re.sub(r'([{\[])\s*,', r'\1', s)
        
        # Fix multiple consecutive commas
        s = re.sub(r',\s*,+', ',', s)
        
        # Step 5: Fix boolean values
        s = re.sub(r':\s*True\b', ': true', s)
        s = re.sub(r':\s*False\b', ': false', s)
        
        # Step 6: Clean up spacing around structural elements
        s = re.sub(r'\s*,\s*', ', ', s)  # Normalize comma spacing
        s = re.sub(r'\s*:\s*', ': ', s)  # Normalize colon spacing
        
        return s
    
    # ---------- Run ----------
    if st.button("🔍 Recommend Candidates"):
        # Process files first to check if we have any
        all_resume_files = process_uploaded_files(uploaded_files, uploaded_zip)
        
        if not all_resume_files or not job_description:
            st.warning("Please enter the job description and upload at least one resume or a ZIP file containing resumes.")
        else:
            job_validation, custom_field_validations = run_pre_evaluation_checks(
                job_title, department, job_description, st.session_state.custom_fields
            )
    
            st.write("**Job Validation:**")
            st.write(job_validation)
            st.write("**Custom Field Validations:**")
            for v in custom_field_validations:
                st.write(v)
    
            EvaluationModel = build_dynamic_model(st.session_state.custom_fields)
            
            # Initialize list to store evaluation results
            if 'evaluations' not in st.session_state:
                st.session_state.evaluations = []
            else:
                st.session_state.evaluations = []  # Clear previous evaluations
    
            with st.spinner("Analyzing resumes with Mistral..."):
                for resume_filename, file_object in all_resume_files:
                    resume_text = extract_text_from_pdf(file_object)
                    prompt = build_eval_prompt(
                        job_title, department, job_description, st.session_state.custom_fields, resume_text
                    )
                    result = call_mistral(prompt)
    
                    st.markdown(f"### 📄 {resume_filename}")
                    if isinstance(result, dict) and "choices" in result:
                        raw_text = result["choices"][0]["message"]["content"]
                        try:
                            cleaned = clean_json_output(raw_text)
                            if not cleaned:
                                st.error("❌ Could not find a JSON object in the model output.")
                                st.write("**Raw Response:**")
                                st.write(raw_text)
                                continue
    
                            try:
                                data = json.loads(cleaned)
                            except json.JSONDecodeError as e:
                                st.error(f"❌ JSON parsing failed: {e}")
                                st.write("**Attempted to clean:**")
                                st.code(cleaned, language="json")
                                st.write("**Raw Response:**")
                                st.write(raw_text)
                                
                                # Try multiple fallback parsing strategies
                                try:
                                    # Strategy 1: More aggressive JSON extraction and cleaning
                                    start = raw_text.find('{')
                                    end = raw_text.rfind('}') + 1
                                    if start != -1 and end > start:
                                        simple_json = raw_text[start:end]
                                        
                                        # Ultra-aggressive cleanup
                                        simple_json = simple_json.replace('"', '"').replace('"', '"')
                                        simple_json = simple_json.replace(''', "'").replace(''', "'")
                                        simple_json = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', simple_json)
                                        simple_json = simple_json.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                                        
                                        # Fix common JSON issues
                                        simple_json = re.sub(r',(\s*[}\]])', r'\1', simple_json)  # Remove trailing commas
                                        simple_json = re.sub(r'([{\[])\s*,', r'\1', simple_json)  # Remove leading commas
                                        simple_json = re.sub(r',\s*,+', ',', simple_json)  # Fix multiple commas
                                        simple_json = re.sub(r'([{,]\s*)([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r'\1"\2":', simple_json)  # Quote field names
                                        simple_json = re.sub(r':\s*True\b', ': true', simple_json)  # Fix booleans
                                        simple_json = re.sub(r':\s*False\b', ': false', simple_json)
                                        simple_json = re.sub(r'\s+', ' ', simple_json)  # Normalize spaces
                                        
                                        # Try to balance braces and brackets
                                        open_braces = simple_json.count('{')
                                        close_braces = simple_json.count('}')
                                        if open_braces > close_braces:
                                            simple_json += '}' * (open_braces - close_braces)
                                        
                                        open_brackets = simple_json.count('[')
                                        close_brackets = simple_json.count(']')
                                        if open_brackets > close_brackets:
                                            simple_json += ']' * (open_brackets - close_brackets)
                                        
                                        data = json.loads(simple_json)
                                        st.info("✅ Enhanced fallback parsing succeeded!")
                                    else:
                                        continue
                                except json.JSONDecodeError as e2:
                                    # Strategy 2: Character-by-character reconstruction
                                    try:
                                        # Find the problematic character around position 1119
                                        error_pos = getattr(e2, 'pos', 1119)
                                        
                                        # Try to fix the specific area around the error
                                        fixed_json = raw_text.replace('"', '"').replace('"', '"')
                                        fixed_json = fixed_json.replace(''', "'").replace(''', "'")
                                        
                                        start = fixed_json.find('{')
                                        end = fixed_json.rfind('}') + 1
                                        if start != -1 and end > start:
                                            fixed_json = fixed_json[start:end]
                                            
                                            # Remove problematic characters around error position
                                            if error_pos < len(fixed_json):
                                                # Look for common issues around the error position
                                                context_start = max(0, error_pos - 50)
                                                context_end = min(len(fixed_json), error_pos + 50)
                                                context = fixed_json[context_start:context_end]
                                                
                                                # Fix common issues in the context
                                                context = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', context)
                                                context = re.sub(r',(\s*[}\]])', r'\1', context)
                                                context = re.sub(r'([{,]\s*)([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r'\1"\2":', context)
                                                
                                                # Reconstruct the JSON
                                                fixed_json = fixed_json[:context_start] + context + fixed_json[context_end:]
                                            
                                            # Final cleanup
                                            fixed_json = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', fixed_json)
                                            fixed_json = re.sub(r',(\s*[}\]])', r'\1', fixed_json)
                                            
                                            data = json.loads(fixed_json)
                                            st.info("✅ Position-specific fix succeeded!")
                                        else:
                                            continue
                                    except Exception as e3:
                                        st.error(f"❌ All parsing strategies failed. Last error: {str(e3)}")
                                        st.write("**Debug info:**")
                                        st.write(f"Original error position: {getattr(e, 'pos', 'unknown')}")
                                        st.write(f"Cleaned JSON length: {len(cleaned) if cleaned else 'N/A'}")
                                        continue
    
                            # Validate with dynamic Pydantic model
                            evaluation = EvaluationModel(**data)
    
                            # ------- UI: core sections (REMOVED experience_relevance) -------
                            col1, col2 = st.columns([1, 2])
                            with col1:
                                st.metric("Key Strengths", f"{evaluation.key_strengths_score}/5")
                                st.caption(f"💭 {evaluation.key_strengths_explanation}")
    
                                st.metric("Experience", f"{evaluation.experience_score}/5")
                                st.caption(f"💭 {evaluation.experience_explanation}")
    
                                st.metric("Skills Match", f"{evaluation.skills_match_score}/5")
                                st.caption(f"💭 {evaluation.skills_match_explanation}")
    
                            with col2:
                                st.write(f"**Candidate:** {evaluation.candidate_name}")
                                st.caption(f"Role: {evaluation.job_title} · Dept: {evaluation.department}")
                                st.write("**💪 Key Strengths**")
                                for s in evaluation.key_strengths:
                                    st.write(f"• {s}")
                                st.write("**⚠️ Potential Concerns**")
                                for c in evaluation.potential_concerns:
                                    st.write(f"• {c}")
    
                            # ------- Custom fields -------
                            if st.session_state.custom_fields:
                                st.write("**📊 Custom Fields**")
                                for field in st.session_state.custom_fields:
                                    fname = field['name']
                                    label = fname.replace('_', ' ').title()
                                    val = getattr(evaluation, fname, None)
                                    sval = getattr(evaluation, f"{fname}_score", None)
                                    expl = getattr(evaluation, f"{fname}_explanation", None)
    
                                    if sval is not None:
                                        st.metric(label, f"{sval}/5")
                                        if val is not None:
                                            st.caption(f"• Value: {val}")
                                        if expl:
                                            st.caption(f"💭 {expl}")
                                    else:
                                        if val is not None:
                                            st.write(f"**{label}:** {val}")
    
                            # ------- How instructions were applied -------
                            if getattr(evaluation, "custom_considerations", None):
                                st.write("**🧠 How Your Instructions Were Applied**")
                                for item in evaluation.custom_considerations:
                                    st.write(
                                        f"- **{item.field}** → "
                                        f"{'APPLIED' if item.applied else 'NOT APPLIED'} | "
                                        f"_Instruction_: {item.instruction} | "
                                        f"_Impact_: {item.impact}"
                                    )
    
                            # ------- OVERALL (LAST): use model's score directly -------
                            st.divider()
                            cols = st.columns([1, 4])
                            with cols[0]:
                                st.metric("Overall Score", f"{evaluation.overall_score}/5")
                            with cols[1]:
                                st.info(f"**Recommendation:** {evaluation.recommendation}")
                                st.caption(f"💭 {evaluation.overall_explanation}")
                                
                            # Store successful evaluation for Excel export
                            # Create a dictionary to store evaluation with metadata instead of modifying the model directly
                            eval_with_metadata = {
                                "evaluation": evaluation,
                                "custom_fields": st.session_state.custom_fields,
                                "resume_filename": resume_filename
                            }
                            st.session_state.evaluations.append(eval_with_metadata)
    
                        except (ValueError, ValidationError) as e:
                            st.error(f"❌ Failed to validate evaluation: {str(e)}")
                            st.write("**Raw Response:**")
                            st.write(raw_text)
                        except Exception as e:
                            st.error(f"❌ Unexpected error: {str(e)}")
                            st.write("**Raw Response:**")
                            st.write(raw_text)
                    else:
                        st.error("❌ Failed to get response from Mistral.")
                        
            # After all evaluations, offer Excel download if we have results
            if st.session_state.evaluations:
                st.divider()
                st.subheader("📊 Export Results")
                
                # Generate Excel report
                today = datetime.datetime.now().strftime("%Y-%m-%d")
                job_title_slug = job_title.lower().replace(" ", "_")
                filename = f"resume_evaluations_{job_title_slug}_{today}.xlsx"
                
                excel_file = create_excel_report(st.session_state.evaluations)
                
                # Display download button
                st.markdown(get_download_link(excel_file, filename), unsafe_allow_html=True)
                st.caption("Export all evaluation results to Excel for offline review")
    
with tab2:
    st.header("Cover Letter AI Detection")
    st.caption("Simple AI vs Human detection for cover letters")
    
    # ---------- File Upload Section ----------
    st.subheader("📤 Upload Cover Letters")
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("**Option 1: Individual Files**")
        uploaded_cover_letters = st.file_uploader(
            "Upload Cover Letters", 
            type=['pdf'],
            accept_multiple_files=True,
            key="cover_letters",
            help="Upload individual PDF cover letters"
        )
    
    with col2:
        st.markdown("**Option 2: ZIP Archive**")
        uploaded_cover_zip = st.file_uploader(
            "Upload ZIP file containing cover letters",
            type=['zip'],
            key="cover_zip",
            help="Upload a ZIP file containing multiple PDF cover letters"
        )
    
    # ---------- Display uploaded files ----------
    all_cover_files = process_cover_letter_files(uploaded_cover_letters, uploaded_cover_zip)
    
    if all_cover_files:
        st.success(f"📁 Ready to analyze {len(all_cover_files)} cover letter(s)")
        with st.expander("View uploaded files"):
            for filename, _ in all_cover_files:
                st.write(f"• {filename}")
    
    # ---------- Analysis Section ----------
    if st.button("🔍 Analyze Cover Letters", key="analyze_covers"):
        if not all_cover_files:
            st.warning("Please upload at least one cover letter file or ZIP archive.")
        else:
            # Initialize session state for cover letter analyses
            if 'cover_analyses' not in st.session_state:
                st.session_state.cover_analyses = []
            else:
                st.session_state.cover_analyses = []  # Clear previous analyses
            
            with st.spinner("Analyzing cover letters for AI detection..."):
                for cover_filename, file_content in all_cover_files:
                    st.markdown(f"### 📄 {cover_filename}")
                    
                    # Extract text from PDF
                    cover_text = extract_text_from_pdf(io.BytesIO(file_content))
                    
                    if not cover_text.strip():
                        st.error("❌ Could not extract text from this file.")
                        continue
                    
                    # Analyze with AI
                    result = analyze_cover_letter_with_ai(cover_text)
                    
                    if result:
                        try:
                            # Clean and parse JSON response
                            cleaned = clean_json_output_cover_letter(result)
                            if not cleaned:
                                st.error("❌ Could not extract analysis from AI response.")
                                st.write("**Raw Response:**")
                                st.write(result)
                                continue
                            
                            # Parse JSON
                            try:
                                data = json.loads(cleaned)
                            except json.JSONDecodeError as e:
                                st.error(f"❌ JSON parsing failed: {e}")
                                st.write("**Attempted to clean:**")
                                st.code(cleaned, language="json")
                                continue
                            
                            # Set filename in data
                            data['file_name'] = cover_filename
                            
                            # Validate with Pydantic model
                            analysis = CoverLetterAnalysis(**data)
                            
                            # Store analysis for Excel export
                            st.session_state.cover_analyses.append(analysis)
                            
                            # ------- Display Results -------
                            col1, col2 = st.columns([1, 1])
                            
                            with col1:
                                # Classification with color coding
                                if analysis.classification == "AI-Generated":
                                    st.error(f"🤖 **{analysis.classification}**")
                                    st.metric("AI Probability", f"{analysis.ai_generated_probability:.1f}%")
                                else:
                                    st.success(f"👤 **{analysis.classification}**")
                                    st.metric("Human Probability", f"{100-analysis.ai_generated_probability:.1f}%")
                                
                                st.metric("Confidence", analysis.confidence_level)
                            
                            with col2:
                                st.write(f"**Applicant:** {analysis.applicant_name}")
                                
                                st.write("**🔍 Key Indicators:**")
                                for indicator in analysis.key_indicators:
                                    st.write(f"• {indicator}")
                            
                            st.divider()
                        
                        except ValidationError as e:
                            st.error(f"❌ Failed to validate analysis: {str(e)}")
                            st.write("**Raw Response:**")
                            st.write(result)
                        except Exception as e:
                            st.error(f"❌ Unexpected error: {str(e)}")
                            st.write("**Raw Response:**")
                            st.write(result)
                    else:
                        st.error("❌ Failed to get response from AI.")
            
            # After all analyses, offer Excel download if we have results
            if st.session_state.cover_analyses:
                st.divider()
                st.subheader("📊 Export Cover Letter Analysis")
                
                # Generate Excel report
                today = datetime.datetime.now().strftime("%Y-%m-%d")
                filename = f"cover_letter_analysis_{today}.xlsx"
                
                excel_file = create_cover_letter_excel_report(st.session_state.cover_analyses)
                
                # Display download button
                st.markdown(get_download_link_cover_letter(excel_file, filename), unsafe_allow_html=True)
                st.caption("Export all cover letter analyses to Excel for offline review")
                
                # Display summary statistics
                ai_count = sum(1 for a in st.session_state.cover_analyses if a.classification == "AI-Generated")
                human_count = len(st.session_state.cover_analyses) - ai_count
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Analyzed", len(st.session_state.cover_analyses))
                with col2:
                    st.metric("AI-Generated", ai_count, delta=f"{ai_count/len(st.session_state.cover_analyses)*100:.1f}%")
                with col3:
                    st.metric("Human-Written", human_count, delta=f"{human_count/len(st.session_state.cover_analyses)*100:.1f}%")